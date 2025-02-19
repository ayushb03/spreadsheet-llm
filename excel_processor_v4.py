import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import load_workbook
import logging
from typing import Dict, List, Union, Optional
from pathlib import Path
import os
import json
import time

# Configure logging
os.makedirs('logs', exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('logs/excel_processor.log')
    ]
)
logger = logging.getLogger(__name__)


class SpreadsheetLLM:
    def __init__(self, spreadsheet: pd.DataFrame):
        """Initialize with a pandas DataFrame."""
        self.spreadsheet = spreadsheet
        self.rows, self.cols = spreadsheet.shape
        logger.info(f"Initialized SpreadsheetLLM with {self.rows} rows and {self.cols} columns")

    def vanilla_encoding(self) -> str:
        """Encode spreadsheet in Markdown-like style."""
        encoded = []
        col_mapping = {col: idx for idx, col in enumerate(self.spreadsheet.columns)}
        
        for (row, col), value in self.spreadsheet.stack().items():
            col_idx = col_mapping[col]
            address = f"{chr(ord('A') + col_idx)}{row + 1}"
            value_str = str(value) if pd.notna(value) else ""
            encoded.append(f"{address},{value_str}")
        
        logger.debug(f"Vanilla encoding completed with {len(encoded)} cells")
        return "|".join(encoded)

    def structural_anchor_extraction(self, k: int = 4) -> pd.DataFrame:
        """
        Filter out empty rows/columns while preserving structure.
        This method identifies structural anchors (heterogeneous rows/columns at table boundaries).
        """
        non_empty_rows = set(self.spreadsheet.index[self.spreadsheet.notna().any(axis=1)])
        non_empty_cols = set(self.spreadsheet.columns[self.spreadsheet.notna().any(axis=0)])

        preserved_rows = set()
        preserved_cols = set()

        # Create a mapping from column names to their numeric indices
        col_indices = {col: idx for idx, col in enumerate(self.spreadsheet.columns)}

        for row in non_empty_rows:
            preserved_rows.update(range(max(0, row - k), min(self.rows, row + k + 1)))
        for col in non_empty_cols:
            col_idx = col_indices[col]  # Get the numeric index of the column
            preserved_cols.update(range(max(0, col_idx - k), min(self.cols, col_idx + k + 1)))

        # Convert preserved_cols back to column names
        col_names = [self.spreadsheet.columns[i] for i in preserved_cols]
        
        # Use loc instead of iloc for label-based indexing
        compact_spreadsheet = self.spreadsheet.loc[
            list(preserved_rows), 
            col_names
        ]
        logger.info(f"Structural extraction preserved {len(preserved_rows)} rows and {len(preserved_cols)} columns")
        return compact_spreadsheet

    def inverted_index_translation(self) -> Dict[str, List[str]]:
        """
        Perform Inverted-index Translation to aggregate cells with identical values.
        """
        inverted_index = defaultdict(list)
        col_mapping = {col: idx for idx, col in enumerate(self.spreadsheet.columns)}

        for (row, col), value in self.spreadsheet.stack().items():
            if pd.notna(value):
                col_idx = col_mapping[col]
                key = str(value)
                inverted_index[key].append(f"{chr(col_idx + 65)}{row + 1}")  # Convert to Excel-style addresses

        logger.info("Inverted-index translation completed")
        return dict(inverted_index)

    def data_format_aggregation(self) -> Dict[str, List[str]]:
        """
        Perform Data-format-aware Aggregation to group cells by their formats.
        """
        format_dict = {
            "IntNum": lambda x: isinstance(x, int),
            "FloatNum": lambda x: isinstance(x, float),
            "Date": lambda x: isinstance(x, str) and "-" in x,
            "Percentage": lambda x: isinstance(x, str) and "%" in x,
            "Others": lambda x: True  # Default category
        }

        aggregated_data = defaultdict(list)
        col_mapping = {col: idx for idx, col in enumerate(self.spreadsheet.columns)}

        for (row, col), value in self.spreadsheet.stack().items():
            if pd.notna(value):
                col_idx = col_mapping[col]
                for fmt, condition in format_dict.items():
                    if condition(value):
                        aggregated_data[fmt].append(f"{chr(col_idx + 65)}{row + 1}")
                        break

        logger.info("Data-format aggregation completed")
        return dict(aggregated_data)

    def compress_spreadsheet(self) -> Dict:
        """
        Compress spreadsheet using all modules: Structural Anchor Extraction, Inverted Index Translation, 
        and Data Format Aggregation.
        """
        logger.info("Starting spreadsheet compression")
        compact_spreadsheet = self.structural_anchor_extraction()
        inverted_index = self.inverted_index_translation()
        aggregated_data = self.data_format_aggregation()
        logger.info("Spreadsheet compression completed")
        return {
            "compact_spreadsheet": compact_spreadsheet,
            "inverted_index": inverted_index,
            "aggregated_data": aggregated_data
        }


class ExcelLoader:
    def __init__(self, file_path: Union[str, Path]):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        logger.info(f"Initialized ExcelLoader with file: {self.file_path}")

    def load_sheets(self) -> Dict[str, pd.DataFrame]:
        """Load all sheets from the Excel file."""
        try:
            excel_data = pd.read_excel(self.file_path, sheet_name=None)
            logger.info(f"Loaded {len(excel_data)} sheets from {self.file_path}")
            return excel_data
        except Exception as e:
            logger.error(f"Error loading Excel file: {str(e)}")
            raise

    def extract_formulas(self) -> Dict[str, Dict[str, str]]:
        """
        Extract formulas from the Excel file using OpenPyXL.
        :return: Dictionary of sheet names and their formulas.
        """
        workbook = load_workbook(self.file_path, data_only=False)
        formulas = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_formulas = {
                f"{cell.coordinate}": cell.value
                for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")
            }
            formulas[sheet_name] = sheet_formulas
        return formulas


class SheetCompressor:
    def __init__(self, sheets: Dict[str, pd.DataFrame]):
        """
        Initialize the compressor with a dictionary of sheets.
        :param sheets: Dictionary of sheet names and their DataFrames.
        """
        self.sheets = sheets

    def compress_all_sheets(self) -> Dict[str, Dict]:
        """
        Compress all sheets using SHEETCOMPRESSOR.
        :return: Dictionary of compressed representations for each sheet.
        """
        compressed_sheets = {}
        for sheet_name, df in self.sheets.items():
            compressor = SpreadsheetLLM(df)
            compressed_data = compressor.compress_spreadsheet()
            compressed_sheets[sheet_name] = compressed_data
        return compressed_sheets


class FormulaAnalyzer:
    def __init__(self, formulas: Dict[str, Dict[str, str]]):
        """
        Initialize the analyzer with extracted formulas.
        :param formulas: Dictionary of sheet names and their formulas.
        """
        self.formulas = formulas

    def analyze_formulas(self) -> Dict[str, Dict[str, List[str]]]:
        """
        Analyze formulas to extract dependencies and semantics.
        :return: Dictionary of analyzed formula relationships.
        """
        analyzed_formulas = {}
        for sheet_name, sheet_formulas in self.formulas.items():
            dependencies = {}
            for cell, formula in sheet_formulas.items():
                deps = [token for token in formula.split() if token.isalnum()]
                dependencies[cell] = deps
            analyzed_formulas[sheet_name] = dependencies
        return analyzed_formulas


class OutputGenerator:
    def __init__(self, compressed_sheets: Dict, formulas: Dict):
        """
        Initialize the output generator.
        :param compressed_sheets: Compressed representations of sheets.
        :param formulas: Analyzed formula relationships.
        """
        self.compressed_sheets = compressed_sheets
        self.formulas = formulas

    def generate_output(self) -> Dict:
        """
        Generate the final structured output.
        :return: Dictionary containing all processed data.
        """
        return {
            "compressed_sheets": self.compressed_sheets,
            "formulas": self.formulas
        }


class ExcelPipeline:
    def __init__(self, file_path: Union[str, Path]):
        self.file_path = Path(file_path)
        # Create output directory based on input file name
        self.output_dir = Path(f"outputs/{self.file_path.stem}")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Outputs will be saved in: {self.output_dir}")

    def run_pipeline(self) -> Dict:
        """Execute the complete processing pipeline."""
        logger.info("Starting Excel processing pipeline")
        try:
            # Step 1: Load Excel file
            loader = ExcelLoader(self.file_path)
            sheets = loader.load_sheets()
            formulas = loader.extract_formulas()

            # Step 2: Compress sheets
            compressor = SheetCompressor(sheets)
            compressed_sheets = compressor.compress_all_sheets()

            # Step 3: Analyze formulas
            formula_analyzer = FormulaAnalyzer(formulas)
            analyzed_formulas = formula_analyzer.analyze_formulas()

            # Step 4: Generate output
            output_generator = OutputGenerator(compressed_sheets, analyzed_formulas)
            output = output_generator.generate_output()

            # Save all outputs
            self._save_outputs(output)
            logger.info("Excel processing pipeline completed successfully")
            return output

        except Exception as e:
            logger.error(f"Error in pipeline execution: {str(e)}", exc_info=True)
            raise

    def _save_outputs(self, output: Dict) -> None:
        """Save all pipeline outputs to files."""
        # Save compressed sheets
        compressed_dir = self.output_dir / "compressed"
        compressed_dir.mkdir(exist_ok=True)
        for sheet_name, data in output['compressed_sheets'].items():
            data['compact_spreadsheet'].to_csv(compressed_dir / f"{sheet_name}_compact.csv")
            with open(compressed_dir / f"{sheet_name}_inverted_index.json", 'w') as f:
                json.dump(data['inverted_index'], f)
            with open(compressed_dir / f"{sheet_name}_aggregated_data.json", 'w') as f:
                json.dump(data['aggregated_data'], f)

        # Save formulas
        formulas_dir = self.output_dir / "formulas"
        formulas_dir.mkdir(exist_ok=True)
        for sheet_name, formula in output['formulas'].items():
            with open(formulas_dir / f"{sheet_name}_formulas.json", 'w') as f:
                json.dump(formula, f)

        logger.info(f"All outputs saved in {self.output_dir}")


if __name__ == "__main__":
    try:
        start_time = time.time()
        file_path = "test-example.xlsx"
        logger.info(f"Starting processing of {file_path}")
        
        pipeline = ExcelPipeline(file_path)
        result = pipeline.run_pipeline()
        
        end_time = time.time()
        total_time = end_time - start_time
        logger.info(f"Processing completed successfully in {total_time:.2f} seconds")
        print(result)
        
    except Exception as e:
        end_time = time.time()
        total_time = end_time - start_time
        logger.critical(f"Critical error after {total_time:.2f} seconds in main execution: {str(e)}", exc_info=True)
        raise